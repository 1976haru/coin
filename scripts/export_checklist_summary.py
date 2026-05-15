"""체크리스트 xlsx 에서 항목 요약을 markdown 으로 출력.

사용:
    python scripts/export_checklist_summary.py [path-to-xlsx]

기본 경로: ../agent_trader_crypto_os_v6_structure_checklist.xlsx
"""
from __future__ import annotations
import sys
import io
from pathlib import Path

# Windows cp949 회피
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


def main(xlsx_path: str | None = None):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 이 필요합니다. pip install openpyxl")
        return 1

    if xlsx_path is None:
        default = Path(__file__).resolve().parent.parent.parent / "agent_trader_crypto_os_v6_structure_checklist.xlsx"
        xlsx_path = str(default)

    if not Path(xlsx_path).exists():
        print(f"파일 없음: {xlsx_path}")
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # 02_통합체크리스트 시트 찾기
    target = None
    for name in wb.sheetnames:
        if "체크리스트" in name or "통합" in name:
            target = name
            break
    if target is None:
        target = wb.sheetnames[2] if len(wb.sheetnames) > 2 else wb.sheetnames[0]

    ws = wb[target]
    print(f"# Checklist Summary — {target}\n")
    print("| # | Phase | 모듈 | 우선순위 | 상태 |")
    print("|---:|---|---|---|---|")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    # 헤더 위치 찾기 (R3 라인부터 데이터 시작 가정)
    for row in rows[3:]:  # skip title rows
        if not row or row[0] is None:
            continue
        try:
            num = int(row[0])
        except Exception:
            continue
        phase = row[1] or ""
        module = row[3] or ""
        priority = row[7] or ""
        status = row[9] or ""
        print(f"| {num} | {phase} | {module} | {priority} | {status} |")

    return 0


if __name__ == "__main__":
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    raise SystemExit(main(arg))
